import openpyxl
from openpyxl import load_workbook
wb = load_workbook('D:\\Infogistic\\Python worksheets\\Book12.xlsx')
print(wb.sheetnames)