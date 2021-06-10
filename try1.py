import openpyxl
import requests
import os
path = "D:\\Infogistic\\Python worksheets\\Test.xlsx"
# loading Excel sheet
wb = openpyxl.load_workbook(path.strip())
# loading active sheet
sheet = wb.active

# Getting URL
for i in range (1, sheet.max_row):
    url_obj = sheet.cell(row=i+1, column=1)
    url_value = url_obj.value
    response = requests.get(url_value)
    q = response.status_code
    print(q)
    print(url_value)
    for j in range (1, sheet.max_row):
        exp_obj = sheet.cell(row=j+1, column=2)
        exp_value = exp_obj.value
        for l in range(1, sheet.max_row):
            code_obj = sheet.cell(row=l+1, column=4)
        if response.status_code == exp_value:
            for k in range(1, sheet.max_row):
                result_obj = sheet.cell(row=k+1, column=3)
                result_obj.value = "Pass"
                code_obj.value = response.status_code
            else:
                for k in range(1, sheet.max_row):
                    result_obj = sheet.cell(row=k+1, column=3)
                    result_obj.value = "Fail"
                code_obj.value = response.status_code


#url_obj = sheet.cell(row=2, column=1)
#url_value = url_obj.value

# Getting access to Result cell
#result_obj = sheet.cell(row=2, column=3)

# getting access to Code cell
#code_obj = sheet.cell(row=2, column=4)

# Getting access to expected code cell
#exp_obj = sheet.cell(row=2, column=2)
#exp_value = exp_obj.value

# Getting response of server
#response = requests.get(url_value)
#if response.status_code == exp_value:
    #result_obj.value = "Pass"
    #code_obj.value = response.status_code
#else:
    #result_obj.value = "Fail"
    #code_obj.value = response.status_code
wb.save("D:\\Infogistic\\Python worksheets\\Results.xlsx")
os.system('start excel.exe "D:\\Infogistic\\Python worksheets\\Results.xlsx"')
