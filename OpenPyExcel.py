import openpyxl
import requests
import os
path = "D:\\Infogistic\\Python worksheets\\Test.xlsx"
# loading Excel sheet
wb = openpyxl.load_workbook(path.strip())
# loading active sheet
sheet = wb.active

# Getting URL
for i in range (1, sheet.max_row+1):
    url_obj = sheet.cell(row=i, column=1)
    print(url_obj.value)
url_obj = sheet.cell(row=2, column=1)
url_value = url_obj.value

# Getting access to Result cell
result_obj = sheet.cell(row=2, column=3)

# getting access to Code cell
code_obj = sheet.cell(row=2, column=4)

# Getting access to expected code cell
exp_obj = sheet.cell(row=2, column=2)
exp_value = exp_obj.value

# Getting response of server
response = requests.get(url_value)
if response.status_code == exp_value:
    result_obj.value = "Pass"
    code_obj.value = response.status_code
else:
    result_obj.value = "Fail"
    code_obj.value = response.status_code
wb.save("D:\\Infogistic\\Python worksheets\\Results.xlsx")
os.system('start excel.exe "D:\\Infogistic\\Python worksheets\\Results.xlsx"')
