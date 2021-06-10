import json
import openpyxl
import requests
import os
import time
print("\033[1;32;40mWorking on Your Report, Please Wait.....")
path = "D:\\Infogistic\\Python worksheets\\Test.xlsx"
# loading Excel sheet
wb = openpyxl.load_workbook(path.strip())
# loading active sheet
sheet = wb.active

# Getting URL
for i in range(1, sheet.max_row):
    type_cell = sheet.cell(row=i+1, column=1)
    type_cell_value = type_cell.value
    if type_cell_value == "Post":
        data_cell = sheet.cell(row=i+1, column=2)
        data_cell_value = data_cell.value
        print(data_cell_value)
        j = json.dumps(data_cell_value)
        a = {
           "email": "eve.holt@reqres.in",
           "password": "cityslicka"
            }
        print(a)
        url_cell = sheet.cell(row=i + 1, column=3)
        url_value = url_cell.value
        response = requests.post(url_value, data=data_cell_value)
        response_content = response.content
        response_code = response.status_code
        exp_code_cell = sheet.cell(row=i + 1, column=4)
        exp_code_value = exp_code_cell.value
        if response_code == exp_code_value:
            result_cell = sheet.cell(row=i + 1, column=5)
            result_cell.value = "Pass"
            actual_code_cell = sheet.cell(row=i + 1, column=6)
            actual_code_cell.value = response_code
            content_cell = sheet.cell(row=i + 1, column=7)
            content_cell.value = response_content
        else:
            result_cell = sheet.cell(row=i + 1, column=5)
            result_cell.value = "Fail"
            actual_code_cell = sheet.cell(row=i + 1, column=6)
            actual_code_cell.value = response_code
            content_cell = sheet.cell(row=i + 1, column=7)
            content_cell.value = response_content
    else:
        url_cell = sheet.cell(row=i + 1, column=3)
        url_value = url_cell.value
        response = requests.get(url_value)
        response_content = response.content
    #print(response_content)
        response_code = response.status_code
        exp_code_cell = sheet.cell(row=i+1, column=4)
        exp_code_value = exp_code_cell.value
        if response_code == exp_code_value:
            result_cell = sheet.cell(row=i+1, column=5)
            result_cell.value = "Pass"
            actual_code_cell = sheet.cell(row=i+1, column=6)
            actual_code_cell.value = response_code
            content_cell = sheet.cell(row=i+1, column=7)
            content_cell.value = response_content
        else:
            result_cell = sheet.cell(row=i + 1, column=5)
            result_cell.value = "Fail"
            actual_code_cell = sheet.cell(row=i + 1, column=6)
            actual_code_cell.value = response_code
            content_cell = sheet.cell(row=i + 1, column=7)
            content_cell.value = response_content
print("Opening the report")
time.sleep(2)
wb.save("D:\\Infogistic\\Python worksheets\\Results.xlsx")
os.system('start excel.exe "D:\\Infogistic\\Python worksheets\\Results.xlsx"')
